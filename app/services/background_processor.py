import os
import logging
import threading
import time
from collections import deque
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from flask import current_app
from app.extensions import db
from app.models.processing_job import ProcessingJob
from app.models.document import Document
from app.models.user import User
from app.models.project import Project
from app.services.document_service import DocumentService

logger = logging.getLogger(__name__)

class BackgroundProcessor:
    """Background processor for handling long-running file processing tasks"""
    
    def __init__(self):
        self.workers = {}  # Track active workers
        self.worker_lock = Lock()
        self.pending_jobs = deque()
        self.pending_job_ids = set()

        # Concurrency: match the number of simultaneous API calls allowed in ai_service.py
        _api_concurrency = int(os.environ.get('API_CONCURRENCY', 5))
        self.max_workers = max(1, int(os.environ.get('MAX_ACTIVE_JOBS', 2)))  # job-level worker threads
        self.max_parallel_files = _api_concurrency  # file-level worker threads

        self.progress_lock = Lock()  # Thread-safe progress updates

        # **ABSOLUTE DUPLICATE PREVENTION SYSTEM**
        self.processing_registry = {}  # {project_id:filename → thread_id}
        self.registry_lock = Lock()

        # Large-scale processing configuration
        self.batch_size = int(os.environ.get('BATCH_SIZE', 100))
        self.checkpoint_frequency = int(os.environ.get('CHECKPOINT_INTERVAL', 50)) // self.batch_size + 1
        self.app = None  # Will be set when needed for context

        # ── Excel write buffer ────────────────────────────────────────────
        # Instead of reading+writing the entire xlsx on every file (O(n²)),
        # we buffer rows in memory and flush in bulk.  Each project gets its
        # own lock so concurrent jobs for different projects don't collide.
        self._excel_buffers: dict = {}       # {excel_path: [row_dict, ...]}
        self._excel_locks:   dict = {}       # {excel_path: threading.Lock}
        self._excel_meta_lock = Lock()       # guards the two dicts above
        self._EXCEL_FLUSH_SIZE = int(os.environ.get('EXCEL_FLUSH_SIZE', 10))

        logger.info(
            f"BackgroundProcessor initialized: api_concurrency={_api_concurrency}, "
            f"max_active_jobs={self.max_workers}, "
            f"batch_size={self.batch_size}, excel_flush={self._EXCEL_FLUSH_SIZE}"
        )
    
    def _get_supported_extensions(self):
        """Get the set of supported file extensions"""
        return {
            'pdf', 'txt', 'doc', 'docx', 'rtf', 'dot', 'dotx', 'hwp', 'hwpx',
            'jpg', 'jpeg', 'png', 'webp', 'bmp', 'gif',
            'xls', 'xlsx', 'csv', 'tsv',
            'mp3', 'wav', 'aiff', 'aac', 'ogg', 'flac',
            'mp4', 'mov', 'avi', 'flv', 'mpg', 'mpeg', 'webm', 'wmv', '3gpp',
            'js', 'html', 'css', 'md', 'yaml', 'yml', 'json', 'xml', 'py', 'cpp', 'java',
            'zip'
        }
    
    def _is_supported_file(self, filename):
        """Check if file is supported"""
        if '.' not in filename:
            return False
        ext = filename.rsplit('.', 1)[-1].lower()
        return ext in self._get_supported_extensions()
    
    def _discover_files_generator(self, folder_path, project_id):
        """Memory-efficient generator for discovering files with ABSOLUTE duplicate prevention"""
        allowed_extensions = self._get_supported_extensions()
        discovered_count = 0
        
        try:
            # Use os.scandir for better performance than listdir
            with os.scandir(folder_path) as entries:
                for entry in entries:
                    if entry.is_file() and self._is_supported_file(entry.name):
                        # **LEVEL 1: Database duplicate check**
                        existing_doc = Document.query.filter_by(
                            project_id=project_id,
                            filename=entry.name
                        ).first()
                        
                        if existing_doc and existing_doc.processed:
                            logger.debug(f"Skipping {entry.name} - already processed in database")
                            continue
                            
                        # **LEVEL 2: Registry-based duplicate prevention**
                        # Check if this file is currently being processed by ANY thread
                        registry_key = f"{project_id}:{entry.name}"
                        with self.registry_lock:
                            if registry_key in self.processing_registry:
                                logger.debug(f"Skipping {entry.name} - currently being processed by thread {self.processing_registry[registry_key]}")
                                continue
                            
                            # Reserve this file for processing by marking it in the registry
                            current_thread_id = threading.get_ident()
                            self.processing_registry[registry_key] = current_thread_id
                            logger.debug(f"Reserved {entry.name} for processing by thread {current_thread_id}")
                            
                        discovered_count += 1
                        file_ext = entry.name.rsplit('.', 1)[-1].lower()
                        
                        yield {
                            'filename': entry.name,
                            'file_path': entry.path,
                            'file_ext': file_ext,
                            'discovered_count': discovered_count,
                            'registry_key': registry_key  # Include registry key for cleanup
                        }
                        
                        # Log progress for very large directories
                        if discovered_count % 10000 == 0:
                            logger.info(f"Discovered {discovered_count} files so far in {folder_path}")
                            
        except Exception as e:
            logger.error(f"Error discovering files in {folder_path}: {e}")
            return
    
    def estimate_folder_cost(self, folder_path, user_id):
        """Estimate pages/files for a folder (credits no longer used for billing)"""
        try:
            total_files = 0
            total_pages = 0
            sample_limit = 1000  # Only sample first 1000 files for cost estimation to avoid memory issues
            
            logger.info(f"Starting cost estimation for folder: {folder_path}")
            
            # Use generator to avoid loading all files into memory
            for file_info in self._discover_files_generator(folder_path, None):  # None for project_id in estimation
                total_files += 1
                filename = file_info['filename']
                file_path = file_info['file_path']
                
                # Estimate pages (PDFs get actual count, others count as 1)
                if filename.lower().endswith('.pdf'):
                    try:
                        pages = DocumentService.get_pdf_page_count(file_path)
                        total_pages += pages
                    except Exception:
                        total_pages += 1  # Fallback to 1 page
                else:
                    total_pages += 1
                
                # For very large folders, sample first 1000 files and estimate
                if total_files >= sample_limit:
                    logger.info(f"Large folder detected ({total_files}+ files), estimating based on sample")
                    # Count remaining files without processing
                    remaining_count = 0
                    try:
                        with os.scandir(folder_path) as entries:
                            remaining_count = sum(1 for entry in entries 
                                                if entry.is_file() and self._is_supported_file(entry.name))
                    except Exception:
                        remaining_count = total_files * 2  # Conservative estimate
                    
                    # Estimate based on sample ratio
                    if total_files > 0:
                        avg_pages_per_file = total_pages / total_files
                        estimated_total_pages = int(remaining_count * avg_pages_per_file)
                        logger.info(f"Estimated {remaining_count} total files, {estimated_total_pages} total pages")
                        return {
                            'total_files': remaining_count,
                            'total_pages': estimated_total_pages,
                            'estimated_credits': int(estimated_total_pages * 1),  # 1 credit per page
                            'success': True,
                            'is_estimate': True  # Flag indicating this is an estimate
                        }
                    break
            
            return {
                'total_files': total_files,
                'total_pages': total_pages,
                'estimated_credits': total_pages,
                'success': True
            }
            
        except Exception as e:
            logger.error(f"Error estimating folder cost: {e}")
            return {
                'success': False,
                'error': str(e)
            }
    
    def estimate_files_cost(self, file_paths, user_id):
        """Estimate pages/files for uploaded files.

        This is informational only now that the app is free to use, so keep it
        intentionally cheap. Counting PDF pages here made the request/queueing
        phase slow before processing even started.
        """
        try:
            total_files = len(file_paths)
            total_pages = total_files
            estimated_credits = total_pages
            
            logger.info(f"File cost estimation complete: {total_files} files, {total_pages} pages")
            
            return {
                'success': True,
                'total_files': total_files,
                'total_pages': total_pages,
                'estimated_credits': estimated_credits
            }
            
        except Exception as e:
            logger.error(f"Error estimating file cost: {e}")
            return {'success': False, 'error': str(e)}
    
    def queue_folder_processing(self, project_id, folder_path, user_id):
        """Queue a folder processing job (no credit checks)"""
        try:
            # Estimate cost first
            cost_estimate = self.estimate_folder_cost(folder_path, user_id)
            if not cost_estimate['success']:
                return {'success': False, 'error': cost_estimate['error']}
            
            # Create processing job
            job = ProcessingJob(
                user_id=user_id,
                project_id=project_id,
                job_type='folder_upload',
                folder_path=folder_path,
                total_files=cost_estimate['total_files'],
                estimated_credits=cost_estimate['estimated_credits'],
                status='queued'
            )
            
            db.session.add(job)
            db.session.commit()
            
            # Start processing in background thread
            self.start_processing_job(job.id)
            
            return {
                'success': True,
                'job_id': job.id,
                'estimated_credits': cost_estimate['estimated_credits'],
                'total_files': cost_estimate['total_files']
            }
            
        except Exception as e:
            logger.error(f"Error queuing folder processing: {e}")
            return {'success': False, 'error': str(e)}
    
    def queue_file_processing(self, project_id, file_paths, user_id):
        """Queue regular uploaded files for background processing with progress tracking (no credit checks)"""
        try:
            # Estimate cost for uploaded files
            cost_estimate = self.estimate_files_cost(file_paths, user_id)
            if not cost_estimate['success']:
                return {'success': False, 'error': cost_estimate['error']}
            
            # Create processing job
            job = ProcessingJob(
                user_id=user_id,
                project_id=project_id,
                job_type='file_upload',
                folder_path=None,  # Not applicable for file uploads
                total_files=cost_estimate['total_files'],
                estimated_credits=cost_estimate['estimated_credits'],
                status='queued',
                result_data={'file_paths': file_paths}  # Store file paths for processing
            )
            
            db.session.add(job)
            db.session.commit()
            
            # Start processing in background thread
            self.start_processing_job(job.id)
            
            return {
                'success': True,
                'job_id': job.id,
                'estimated_credits': cost_estimate['estimated_credits'],
                'total_files': cost_estimate['total_files']
            }
            
        except Exception as e:
            logger.error(f"Error queuing file processing: {e}")
            return {'success': False, 'error': str(e)}
    
    def start_processing_job(self, job_id):
        """Start a job immediately or enqueue it for later."""
        with self.worker_lock:
            if job_id in self.workers or job_id in self.pending_job_ids:
                return job_id in self.workers

            if len(self.workers) >= self.max_workers:
                self.pending_jobs.append(job_id)
                self.pending_job_ids.add(job_id)
                logger.info(f"Queued job {job_id}; active workers at capacity ({self.max_workers})")
                return False

            self._launch_worker_locked(job_id)
            return True

    def _launch_worker_locked(self, job_id):
        """Launch a worker thread. Caller must hold worker_lock."""
        worker_thread = threading.Thread(
            target=self._process_job_worker,
            args=(job_id,),
            daemon=True,
            name=f"ProcessingWorker-{job_id}"
        )
        self.workers[job_id] = {
            'thread': worker_thread,
            'started_at': datetime.utcnow()
        }
        worker_thread.start()
        logger.info(f"Started background processing for job {job_id}")

    def _drain_pending_jobs_locked(self):
        """Start queued jobs until capacity is full. Caller must hold worker_lock."""
        while self.pending_jobs and len(self.workers) < self.max_workers:
            next_job_id = self.pending_jobs.popleft()
            self.pending_job_ids.discard(next_job_id)
            self._launch_worker_locked(next_job_id)
    
    def _process_single_file(self, file_info):
        """Process a single file in a thread-safe manner with database retry logic"""
        from app import create_app
        app = create_app()
        
        filename, file_path, file_ext, job_id, project_id = file_info
        
        with app.app_context():
            try:
                # Retry logic for database operations
                max_retries = 3
                retry_delay = 0.5
                
                for attempt in range(max_retries):
                    try:
                        # Start fresh database session for each attempt
                        db.session.rollback()  # Clear any pending transactions
                        
                        # Get fresh instances for thread safety
                        job = ProcessingJob.query.get(job_id)
                        if not job or job.status == 'cancelled':
                            return {'filename': filename, 'status': 'cancelled', 'message': 'Job cancelled'}
                        
                        project = job.project
                        user = User.query.get(job.user_id)
                        
                        # Check for duplicates
                        existing_doc = Document.query.filter_by(
                            project_id=project_id,
                            filename=filename
                        ).first()
                        
                        if existing_doc and existing_doc.processed:
                            logger.info(f"Job {job_id}: Skipping {filename} - already processed")
                            with self.progress_lock:
                                job.update_progress(skipped=1)
                            return {'filename': filename, 'status': 'skipped', 'message': 'Already processed'}
                        
                        # Calculate pages for this file (credits no longer used for billing)
                        if file_ext == 'pdf':
                            try:
                                page_count = DocumentService.get_pdf_page_count(file_path)
                            except Exception:
                                page_count = 1
                        else:
                            page_count = 1
                        
                        # Skip file copying - process directly from original location
                        # This eliminates I/O overhead and improves performance
                        
                        # Get file type category
                        file_type = DocumentService._get_file_type_category(file_ext)
                        
                        # Create Document record with original file path for direct processing
                        document = Document(
                            project_id=project_id,
                            filename=filename,
                            file_path=file_path,  # Use original path, not copied
                            file_type=file_type,
                            page_count=page_count
                        )
                        db.session.add(document)
                        db.session.flush()
                        
                        # Process with AI (this doesn't require database transactions)
                        success = DocumentService.process_document(document.id)
                        
                        if success:
                            logger.info(f"Job {job_id}: Successfully processed {filename}")
                            # Credits usage removed – always report 0
                            return {'filename': filename, 'status': 'success', 'credits_used': 0}
                        else:
                            logger.error(f"Job {job_id}: Failed to process {filename}")
                            return {'filename': filename, 'status': 'failed', 'message': 'AI processing failed'}
                            
                        # If we get here, the operation succeeded, break out of retry loop
                        break
                        
                    except Exception as db_error:
                        db.session.rollback()  # Always rollback on error
                        if "database is locked" in str(db_error) and attempt < max_retries - 1:
                            logger.warning(f"Job {job_id}: Database locked on attempt {attempt + 1}, retrying in {retry_delay}s...")
                            time.sleep(retry_delay)
                            retry_delay *= 2  # Exponential backoff
                            continue
                        else:
                            # Last attempt or non-lock error, log and return error
                            logger.error(f"Job {job_id}: Database error after {max_retries} attempts for {filename}: {db_error}")
                            return {'filename': filename, 'status': 'error', 'message': f'Database error: {str(db_error)}'}
                    
            except Exception as e:
                logger.error(f"Job {job_id}: Error processing {filename}: {e}")
                with self.progress_lock:
                    job = ProcessingJob.query.get(job_id)
                    if job:
                        job.update_progress(failed=1)
                        db.session.rollback()
                return {'filename': filename, 'status': 'error', 'message': str(e)}
    
    def _process_single_file_direct(self, filename, file_path, file_ext, project_id):
        """Process a single file directly to AI without any copying or intermediate storage"""
        if not self.app:
            from app import create_app
            self.app = create_app()
        
        with self.app.app_context():
            try:
                # Check if file exists
                if not os.path.exists(file_path):
                    return {'success': False, 'error': 'File not found', 'credits_used': 0}
                
                # Get project for AI processing fields
                from app.models import Project
                project = Project.query.get(project_id)
                if not project:
                    return {'success': False, 'error': 'Project not found', 'credits_used': 0}
                user = User.query.get(project.user_id)

                # Ensure AI provider is configured for this user before doing work
                from app.services.ai_service import AIService
                ok, msg = AIService.ensure_provider_configured(user=user)
                if not ok:
                    return {'success': False, 'error': msg or 'AI provider not configured', 'credits_used': 0}
                
                # Calculate pages for this file (credits no longer used for billing)
                if file_ext == 'pdf':
                    try:
                        page_count = DocumentService.get_pdf_page_count(file_path)
                    except Exception:
                        page_count = 1
                else:
                    page_count = 1
                
                # Get file type category
                file_type = DocumentService._get_file_type_category(file_ext)
                
                # **DIRECT AI PROCESSING** - Send file directly to AI without database records or copying
                logger.info(f"Processing {filename} directly with AI (no copying, no intermediate storage) - File type: {file_type}")
                
                # Special logging for image files to help debug the issue
                if file_type == 'image':
                    logger.info(f"IMAGE PROCESSING: {filename} (extension: {file_ext}, size: {os.path.getsize(file_path) if os.path.exists(file_path) else 'MISSING'} bytes)")
                
                # Extract data directly from file using AI (use project owner's AI settings)
                extracted_data_list = AIService.extract_data_with_ai(
                    file_path,  # Send file directly from original location
                    project.fields,
                    file_type,
                    user=user
                )
                
                if not extracted_data_list:
                    return {'success': False, 'error': 'AI extraction failed - no data returned', 'credits_used': 0}
                
                # If it's a single dict (image), wrap in a list for consistency
                if isinstance(extracted_data_list, dict):
                    extracted_data_list = [extracted_data_list]

                first_item = extracted_data_list[0] if extracted_data_list else {}
                if isinstance(first_item, dict) and first_item.get('error'):
                    return {
                        'success': False,
                        'error': first_item.get('error', 'AI extraction failed'),
                        'credits_used': 0
                    }
                
                # Merge all page dicts into one: for each field, join values with newlines
                merged_data = {}
                for field in project.fields:
                    field_name = field['name']
                    values = [page_data.get(field_name, '') for page_data in extracted_data_list if page_data.get(field_name)]
                    values = [v for v in values if v and v.lower() != "not found"]
                    merged_data[field_name] = '\n'.join(values) if values else "Not found"
                
                # **FINAL DUPLICATE CHECK** - Double-check for duplicates before database insertion
                # This is a final safety net against race conditions
                final_duplicate_check = Document.query.filter_by(
                    project_id=project_id,
                    filename=filename
                ).first()
                
                if final_duplicate_check and final_duplicate_check.processed:
                    logger.info(f"Final duplicate check: {filename} was already processed by another thread")
                    return {'success': False, 'error': 'Duplicate processed by another thread', 'credits_used': 0}
                
                # **MINIMAL DATABASE OPERATIONS** - Only create document and data record after successful AI processing
                # This ensures we don't waste database records on failed AI processing
                document = Document(
                    project_id=project_id,
                    filename=filename,
                    file_path=file_path,  # Keep original path (no copying)
                    file_type=file_type,
                    page_count=page_count,
                    processed=True  # Mark as processed since AI extraction succeeded
                )
                db.session.add(document)
                db.session.flush()  # Get document ID
                
                # Create data record without storing data in database
                # Data is stored in Excel file only for better performance
                from app.models.document import DataRecord
                data_record = DataRecord(
                    document_id=document.id
                    # data column removed - data stored in Excel file only
                )
                db.session.add(data_record)
                
                # Update Excel file with new data
                self._update_excel_file_direct(project, merged_data, filename)
                
                db.session.commit()
                logger.info(f"Successfully processed {filename} directly with AI")
                
                # Credits usage removed – always report 0
                return {'success': True, 'credits_used': 0}
                    
            except Exception as e:
                db.session.rollback()
                logger.error(f"Error processing {filename} directly: {e}", exc_info=True)
                return {'success': False, 'error': str(e), 'credits_used': 0}
    
    # ── Excel helpers ──────────────────────────────────────────────────────

    def _get_excel_lock(self, excel_path: str) -> Lock:
        """Return (creating if necessary) the per-file write lock."""
        with self._excel_meta_lock:
            if excel_path not in self._excel_locks:
                self._excel_locks[excel_path] = Lock()
                self._excel_buffers[excel_path] = []
            return self._excel_locks[excel_path]

    def _update_excel_file_direct(self, project, merged_data: dict, filename: str):
        """Buffer one row and flush to disk when the buffer is full.

        Uses openpyxl in append mode instead of pandas read-entire-file-write,
        cutting Excel I/O from O(n²) to O(1) per file (amortised O(n/flush_size)).
        """
        try:
            excel_path = os.path.join(project.storage_path, "extracted_data.xlsx")
            row_data = {**merged_data, 'filename': filename,
                        'extracted_date': datetime.utcnow().isoformat()}

            lock = self._get_excel_lock(excel_path)
            with lock:
                self._excel_buffers[excel_path].append(row_data)
                if len(self._excel_buffers[excel_path]) >= self._EXCEL_FLUSH_SIZE:
                    self._flush_excel_buffer(excel_path, project.fields)
        except Exception as e:
            logger.error(f"Error buffering Excel row for {filename}: {e}", exc_info=True)

    def _flush_excel_buffer(self, excel_path: str, fields: list):
        """Write all buffered rows to the xlsx file using openpyxl (called under lock)."""
        rows = self._excel_buffers.get(excel_path, [])
        if not rows:
            return
        try:
            from openpyxl import load_workbook, Workbook

            # Column order: user fields first, then metadata
            col_names = [f['name'] for f in fields] + ['filename', 'extracted_date']

            if os.path.exists(excel_path):
                try:
                    wb = load_workbook(excel_path)
                    ws = wb.active
                except Exception:
                    wb = Workbook()
                    ws = wb.active
                    ws.append(col_names)
            else:
                wb = Workbook()
                ws = wb.active
                ws.append(col_names)

            for row in rows:
                ws.append([str(row.get(c, '')) for c in col_names])

            os.makedirs(os.path.dirname(excel_path), exist_ok=True)
            wb.save(excel_path)
            self._excel_buffers[excel_path] = []
            logger.debug(f"Flushed {len(rows)} rows → {excel_path}")
        except Exception as e:
            logger.error(f"Error flushing Excel buffer to {excel_path}: {e}", exc_info=True)

    def _flush_all_excel_buffers_for_project(self, project):
        """Flush any remaining buffered rows at job completion."""
        excel_path = os.path.join(project.storage_path, "extracted_data.xlsx")
        lock = self._get_excel_lock(excel_path)
        with lock:
            self._flush_excel_buffer(excel_path, project.fields)
    
    def _process_job_worker(self, job_id):
        """Worker thread for processing a job"""
        from app import create_app
        app = create_app()
        
        with app.app_context():
            try:
                # Get job from database
                job = ProcessingJob.query.get(job_id)
                if not job:
                    logger.error(f"Job {job_id} not found")
                    return
                
                job.start()
                logger.info(f"Worker started for job {job_id}")
                
                # Process based on job type
                if job.job_type == 'folder_upload':
                    self._process_folder_job(job)
                elif job.job_type == 'file_upload':
                    self._process_file_upload_job(job)
                else:
                    logger.error(f"Unknown job type: {job.job_type}")
                    job.fail(f"Unknown job type: {job.job_type}")
                
            except Exception as e:
                logger.error(f"Worker error for job {job_id}: {e}")
                job = ProcessingJob.query.get(job_id)
                if job:
                    job.fail(str(e))
            finally:
                # Clean up worker
                with self.worker_lock:
                    self.workers.pop(job_id, None)
                    self._drain_pending_jobs_locked()
                logger.info(f"Worker finished for job {job_id}")
    
    def _process_folder_job(self, job):
        """Process a folder upload job with scalable batch processing for millions of files"""
        if not self.app:
            from app import create_app
            self.app = create_app()
            
        try:
            with self.app.app_context():
                folder_path = job.folder_path
                if not os.path.exists(folder_path):
                    job.complete_with_error("Folder not found")
                    return
                
                logger.info(f"Starting scalable folder processing for job {job.id}: {folder_path}")
                
                # Initialize counters
                total_processed = 0
                total_failed = 0
                total_skipped = 0
                total_credits_used = 0
                
                # Process files in batches to handle unlimited number of files
                current_batch = []
                checkpoint_counter = 0
                
                for file_info in self._discover_files_generator(folder_path, job.project_id):
                    file_tuple = (
                        file_info['filename'], 
                        file_info['file_path'], 
                        file_info['file_ext'],
                        file_info['registry_key']  # Include registry key for cleanup
                    )
                    current_batch.append(file_tuple)
                    
                    # Process batch when it reaches batch_size
                    if len(current_batch) >= self.batch_size:
                        batch_results = self._process_file_batch_scalable(current_batch, job.project_id)
                        
                        # Update counters
                        total_processed += batch_results['processed']
                        total_failed += batch_results['failed']
                        total_skipped += batch_results['skipped']
                        total_credits_used += batch_results['credits_used']
                        
                        # Update progress with absolute values
                        job.update_progress(
                            processed=total_processed,
                            failed=total_failed, 
                            skipped=total_skipped,
                            credits_used=total_credits_used,
                            absolute=True
                        )
                        
                        # No credit deduction – processing is always allowed
                        
                        # Checkpoint every few batches
                        checkpoint_counter += 1
                        if checkpoint_counter >= self.checkpoint_frequency:
                            self._checkpoint_progress(job, total_processed, total_failed, total_skipped, total_credits_used)
                            checkpoint_counter = 0
                        
                        # Clear batch for next iteration
                        current_batch = []
                        
                        # Check if job was cancelled
                        if job.status == 'cancelled':
                            logger.info(f"Job {job.id} was cancelled during processing")
                            return
                
                # Process remaining files in the last batch
                if current_batch:
                    batch_results = self._process_file_batch_scalable(current_batch, job.project_id)
                    
                    total_processed += batch_results['processed']
                    total_failed += batch_results['failed']
                    total_skipped += batch_results['skipped']
                    total_credits_used += batch_results['credits_used']
                    
                    # Update progress after final batch
                    job.update_progress(
                        processed=total_processed,
                        failed=total_failed, 
                        skipped=total_skipped,
                        credits_used=total_credits_used,
                        absolute=True
                    )
                    
                    # No final credit deduction – processing is always allowed
                
                # Final completion
                with self.app.app_context():
                    self._complete_folder_processing(job, total_processed, total_failed, total_skipped, total_credits_used)
                
        except Exception as e:
            logger.error(f"Error in folder processing job {job.id}: {e}", exc_info=True)
            with self.app.app_context():
                job.complete_with_error(str(e))
    
    def _process_file_batch(self, file_batch):
        """Process a batch of files in parallel and return results"""
        processed_files = []
        skipped_files = []
        failed_files = []
        total_credits_used = 0
        
        with ThreadPoolExecutor(max_workers=self.max_parallel_files) as executor:
            # Submit all files in batch for processing
            future_to_file = {
                executor.submit(self._process_single_file, file_info): file_info[0] 
                for file_info in file_batch
            }
            
            # Process completed results
            for future in as_completed(future_to_file):
                filename = future_to_file[future]
                
                try:
                    result = future.result()
                    
                    if result['status'] == 'success':
                        processed_files.append(filename)
                        total_credits_used += result.get('credits_used', 0)
                    elif result['status'] == 'skipped':
                        skipped_files.append(filename)
                    elif result['status'] == 'insufficient_credits':
                        logger.info(f"Batch processing: Insufficient credits at {filename}")
                        failed_files.append(filename)
                    elif result['status'] == 'cancelled':
                        logger.info(f"Batch processing: Cancelled at {filename}")
                        failed_files.append(filename)
                    else:
                        failed_files.append(filename)
                        
                except Exception as e:
                    logger.error(f"Batch processing: Future exception for {filename}: {e}")
                    failed_files.append(filename)
        
        return {
            'processed': processed_files,
            'skipped': skipped_files,
            'failed': failed_files,
            'credits_used': total_credits_used
        }
    
    def _process_file_batch_scalable(self, file_batch, project_id):
        """Process a batch of files in parallel and return aggregated results"""
        if not file_batch:
            return {'processed': 0, 'failed': 0, 'skipped': 0, 'credits_used': 0}
        
        logger.info(f"Processing batch of {len(file_batch)} files")
        
        # Initialize results
        batch_results = {
            'processed': 0,
            'failed': 0,
            'skipped': 0,
            'credits_used': 0
        }
        
        # **REGISTRY-BASED PROCESSING**: Files are already reserved in the registry, no need for additional filtering
        # All files in this batch are guaranteed to be unique and reserved for processing
        files_to_process = []
        for file_data in file_batch:
            if len(file_data) == 4:  # New format with registry key
                filename, file_path, file_ext, registry_key = file_data
            else:  # Fallback for old format (shouldn't happen with new system)
                filename, file_path, file_ext = file_data
                registry_key = f"{project_id}:{filename}"
                logger.warning(f"Using fallback registry key for {filename}")
            
            files_to_process.append((filename, file_path, file_ext, registry_key))
        
        # Process registry-reserved files with cleanup
        with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
            future_to_file = {}
            for filename, file_path, file_ext, registry_key in files_to_process:
                future = executor.submit(self._process_single_file_direct, filename, file_path, file_ext, project_id)
                future_to_file[future] = (filename, registry_key)
            
            # Collect results as they complete
            for future in as_completed(future_to_file):
                filename, registry_key = future_to_file[future]
                try:
                    result = future.result()
                    if result['success']:
                        batch_results['processed'] += 1
                        batch_results['credits_used'] += result.get('credits_used', 0)
                        logger.debug(f"Successfully processed {filename}")
                    else:
                        batch_results['failed'] += 1
                        logger.error(f"Failed to process {filename}: {result.get('error', 'Unknown error')}")
                        
                except Exception as e:
                    batch_results['failed'] += 1
                    logger.error(f"Exception processing {filename}: {e}", exc_info=True)
                finally:
                    # **CLEANUP REGISTRY**: Always remove from processing registry after completion
                    with self.registry_lock:
                        if registry_key in self.processing_registry:
                            del self.processing_registry[registry_key]
                            logger.debug(f"Cleaned up registry entry for {filename}")
                        else:
                            logger.warning(f"Registry key {registry_key} not found during cleanup")
        
        logger.info(f"Batch completed: {batch_results['processed']} processed, {batch_results['failed']} failed")
        return batch_results
    
    def _checkpoint_progress(self, job, processed, failed, skipped, credits_used):
        """Save progress checkpoint during large-scale processing"""
        try:
            logger.info(f"Checkpoint - Job {job.id}: {processed} processed, {failed} failed, {skipped} skipped")
            
            # Update job progress with absolute values
            job.update_progress(
                processed=processed,
                failed=failed,
                skipped=skipped, 
                credits_used=credits_used,
                absolute=True
            )
            
            # Log memory usage for monitoring
            try:
                import psutil
                process = psutil.Process()
                memory_info = process.memory_info()
                logger.info(f"Memory usage - RSS: {memory_info.rss / 1024 / 1024:.1f} MB, VMS: {memory_info.vms / 1024 / 1024:.1f} MB")
            except ImportError:
                pass  # psutil not available
            
        except Exception as e:
            logger.error(f"Error during checkpoint for job {job.id}: {e}", exc_info=True)
    
    def _final_checkpoint(self, job, processed, failed, skipped, credits_used, processed_list, skipped_list, failed_list):
        """Final checkpoint with job completion"""
        with self.progress_lock:
            try:
                # Update final job state (no credit deductions – app is free to use)
                job.processed_files = processed
                job.failed_files = failed  
                job.skipped_files = skipped
                job.credits_used = credits_used
                job.status = 'completed'
                job.completed_at = datetime.utcnow()
                job.result_data = {
                    'processed_files': processed_list[-100:] if len(processed_list) > 100 else processed_list,  # Keep last 100 for UI
                    'skipped_files': skipped_list[-100:] if len(skipped_list) > 100 else skipped_list,
                    'failed_files': failed_list[-100:] if len(failed_list) > 100 else failed_list,
                    'total_batches_processed': len(processed_list) + len(skipped_list) + len(failed_list)
                }
                
                db.session.commit()
                logger.info(f"Job {job.id}: Final checkpoint completed")
                
            except Exception as e:
                logger.error(f"Job {job.id}: Error in final checkpoint: {e}")
                db.session.rollback()
    
    def _complete_folder_processing(self, job, processed, failed, skipped, credits_used):
        """Complete folder processing with final statistics and cleanup"""
        try:
            logger.info(f"Completing folder processing for job {job.id}")

            # Flush any remaining buffered Excel rows before marking complete
            try:
                project = job.project
                if project:
                    self._flush_all_excel_buffers_for_project(project)
            except Exception as flush_err:
                logger.warning(f"Excel flush at folder completion: {flush_err}")

            # Final progress update with absolute values
            job.update_progress(
                processed=processed,
                failed=failed,
                skipped=skipped,
                credits_used=credits_used,
                absolute=True
            )
            
            # Credit usage tracking removed – no billing or balance records
            
            # Generate completion summary
            total_files = processed + failed + skipped
            success_rate = (processed / total_files * 100) if total_files > 0 else 0
            
            summary = f"Folder processing completed: {processed}/{total_files} files processed successfully ({success_rate:.1f}% success rate). Credits used: {credits_used}"
            
            # Mark job as completed - refresh job from database to avoid stale session issues
            try:
                # Update in the same transaction
                job.status = 'completed'
                job.completed_at = datetime.utcnow()
                job.result_summary = summary
                
                # Ensure final counts are set correctly
                job.processed_files = processed
                job.failed_files = failed  
                job.skipped_files = skipped
                job.credits_used = credits_used
                
                # Force a flush to the database
                db.session.flush()
                db.session.commit()
                
                # Double-check completion status in a separate transaction
                fresh_job = ProcessingJob.query.get(job.id)
                if fresh_job and fresh_job.status != 'completed':
                    logger.warning(f"Job {job.id} status inconsistency detected, forcing completion")
                    fresh_job.status = 'completed'
                    fresh_job.completed_at = datetime.utcnow()
                    db.session.commit()
                
                logger.info(f"Successfully updated job {job.id} status to completed with final stats: {processed}/{job.total_files} processed")
            except Exception as e:
                logger.error(f"Error marking job {job.id} as completed: {e}")
                # Final attempt to mark as completed
                try:
                    db.session.rollback()
                    fresh_job = ProcessingJob.query.get(job.id)
                    if fresh_job:
                        fresh_job.status = 'completed'
                        fresh_job.completed_at = datetime.utcnow()
                        db.session.commit()
                except Exception:
                    logger.error(f"Critical failure marking job {job.id} as completed")
            
            logger.info(f"Job {job.id} completed: {summary}")
            
        except Exception as e:
            logger.error(f"Error completing folder processing for job {job.id}: {e}", exc_info=True)
            job.complete_with_error(f"Error in completion: {str(e)}")
    
    def _process_file_upload_job(self, job):
        """Process uploaded files in parallel (same fast path as folder jobs).

        Uses ThreadPoolExecutor + _process_single_file_direct for parallel API
        calls and buffered Excel writes instead of sequential DocumentService path.
        """
        if not self.app:
            from app import create_app
            self.app = create_app()

        try:
            with self.app.app_context():
                file_paths = job.result_data.get('file_paths', [])
                if not file_paths:
                    job.complete_with_error("No files to process")
                    return

                project = Project.query.get(job.project_id)
                if not project:
                    job.complete_with_error("Project not found")
                    return

                logger.info(f"Starting file upload processing for job {job.id}: {len(file_paths)} files (parallel)")

                files_to_process = []
                for item in file_paths:
                    if isinstance(item, dict):
                        file_path = item.get('file_path') or ''
                        filename = item.get('filename') or os.path.basename(file_path)
                    else:
                        file_path = item
                        filename = os.path.basename(file_path)

                    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
                    if self._is_supported_file(filename) and file_path:
                        files_to_process.append((filename, file_path, ext))

                total_processed = 0
                total_failed = 0
                total_skipped = 0
                total_credits_used = 0

                with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                    future_to_file = {
                        executor.submit(
                            self._process_single_file_direct,
                            filename, file_path, file_ext, project.id
                        ): filename
                        for filename, file_path, file_ext in files_to_process
                    }

                    for future in as_completed(future_to_file):
                        fn = future_to_file[future]
                        try:
                            result = future.result()
                            if result['success']:
                                total_processed += 1
                                total_credits_used += result.get('credits_used', 0)
                            else:
                                err = result.get('error', '')
                                if 'already processed' in err.lower() or 'duplicate' in err.lower():
                                    total_skipped += 1
                                else:
                                    total_failed += 1
                                    logger.error(f"Failed {fn}: {err}")
                        except Exception as e:
                            total_failed += 1
                            logger.error(f"Error processing {fn}: {e}")

                        job.update_progress(
                            processed=total_processed,
                            failed=total_failed,
                            skipped=total_skipped,
                            credits_used=total_credits_used,
                            absolute=True
                        )

                self._complete_file_processing(job, total_processed, total_failed, total_skipped, total_credits_used)
                
        except Exception as e:
            logger.error(f"Error in file upload processing for job {job.id}: {e}", exc_info=True)
            job.complete_with_error(f"Processing error: {str(e)}")
    
    def _complete_file_processing(self, job, processed, failed, skipped, credits_used):
        """Complete file processing with final statistics"""
        try:
            logger.info(f"Completing file processing for job {job.id}")

            # Flush any remaining buffered Excel rows
            try:
                project = job.project
                if project:
                    self._flush_all_excel_buffers_for_project(project)
            except Exception as flush_err:
                logger.warning(f"Excel flush at file completion: {flush_err}")
            
            # Create summary message
            total_files = processed + failed + skipped
            summary_parts = []
            if processed > 0:
                summary_parts.append(f"{processed} processed")
            if failed > 0:
                summary_parts.append(f"{failed} failed")
            if skipped > 0:
                summary_parts.append(f"{skipped} skipped")
            
            summary = f"File upload complete: {', '.join(summary_parts)} ({total_files} total files)"
            
            # Final progress update with absolute values 
            job.update_progress(
                processed=processed,
                failed=failed,
                skipped=skipped,
                credits_used=credits_used,
                absolute=True
            )
            
            # Mark job as completed with preserved final values
            try:
                # Update in the same transaction
                job.status = 'completed'
                job.completed_at = datetime.utcnow()
                job.result_summary = summary
                
                # Ensure final counts are preserved
                job.processed_files = processed
                job.failed_files = failed
                job.skipped_files = skipped
                job.credits_used = credits_used
                
                # Force a flush to ensure database is updated
                db.session.flush()
                db.session.commit()
                
                # Double-check completion status
                fresh_job = ProcessingJob.query.get(job.id)
                if fresh_job and fresh_job.status != 'completed':
                    logger.warning(f"Job {job.id} status inconsistency detected, forcing completion")
                    fresh_job.status = 'completed'
                    fresh_job.completed_at = datetime.utcnow()
                    db.session.commit()
                
                logger.info(f"File upload job {job.id} completed: {summary}")
            except Exception as e:
                logger.error(f"Error marking job {job.id} as completed: {e}")
                # Final attempt to mark as completed
                try:
                    db.session.rollback()
                    fresh_job = ProcessingJob.query.get(job.id)
                    if fresh_job:
                        fresh_job.status = 'completed'
                        fresh_job.completed_at = datetime.utcnow()
                        db.session.commit()
                except Exception:
                    logger.error(f"Critical failure marking job {job.id} as completed")
                
        except Exception as e:
            logger.error(f"Error completing file processing for job {job.id}: {e}", exc_info=True)
            job.complete_with_error(f"Error in completion: {str(e)}")
    
    def get_job_status(self, job_id):
        """Get the status of a processing job"""
        job = ProcessingJob.query.get(job_id)
        if not job:
            return {'error': 'Job not found'}
        
        # Get remaining credits for user
        user = User.query.get(job.user_id)
        remaining_credits = user.credits if user else 0
        
        # For backward compatibility, include both old and new field names
        return {
            'job_id': job.id,
            'status': job.status,
            'progress_percentage': job.get_progress_percentage(),
            'processed': job.processed_files,  # For backward compatibility
            'processed_files': job.processed_files,
            'failed': job.failed_files,  # For backward compatibility
            'failed_files': job.failed_files,
            'skipped': job.skipped_files,  # For backward compatibility
            'skipped_files': job.skipped_files,
            'total_files': job.total_files,
            'credits_used': job.credits_used,
            'used_credits': job.credits_used,  # For backward compatibility
            'estimated_credits': job.estimated_credits,
            'remaining_credits': remaining_credits,  # Add remaining credits for UI
            'credits_remaining': remaining_credits,  # For backward compatibility
            'estimated_time_remaining': job.get_estimated_time_remaining(),
            'created_at': job.created_at.isoformat(),
            'started_at': job.started_at.isoformat() if job.started_at else None,
            'completed_at': job.completed_at.isoformat() if job.completed_at else None,
            'error_message': job.error_message
        }
    
    def cancel_job(self, job_id, user_id):
        """Cancel a processing job"""
        job = ProcessingJob.query.filter_by(id=job_id, user_id=user_id).first()
        if not job:
            return {'error': 'Job not found or access denied'}
        
        if job.status in ['completed', 'failed', 'cancelled']:
            return {'error': 'Job cannot be cancelled (already finished)'}
        
        job.cancel("Cancelled by user")
        return {'success': True, 'message': 'Job cancelled successfully'}
        
    def cleanup_processing_registry(self, project_id=None):
        """Clean up the processing registry (useful for system recovery)"""
        with self.registry_lock:
            if project_id:
                # Clean up entries for a specific project
                keys_to_remove = [key for key in self.processing_registry.keys() if key.startswith(f"{project_id}:")]
                for key in keys_to_remove:
                    del self.processing_registry[key]
                logger.info(f"Cleaned up {len(keys_to_remove)} registry entries for project {project_id}")
            else:
                # Clean up all entries
                count = len(self.processing_registry)
                self.processing_registry.clear()
                logger.info(f"Cleaned up {count} total registry entries")
                
    def get_processing_registry_status(self):
        """Get current status of the processing registry"""
        with self.registry_lock:
            return {
                'total_entries': len(self.processing_registry),
                'entries_by_project': {},
                'active_threads': list(set(self.processing_registry.values()))
            }

# Global instance
background_processor = BackgroundProcessor()